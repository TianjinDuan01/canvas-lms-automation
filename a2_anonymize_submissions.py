"""
A2 Game Submission Anonymizer
==============================
Step 1 — Validate filenames (run before deadline, no download)
  Checks that each student submitted files named correctly per Public_Assigned_Names.xlsx.
  Outputs a report listing students to reach out to.

Step 2 — Download and anonymize HTML files (run after Step 1 is clean)
  Downloads each student's latest HTML submission and renames it
  according to the fixed mapping in PRIVATE_Mapping.xlsx.
  Output folder is ready to upload to the professor's website.

Usage:
  pip install requests openpyxl
  python a2_anonymize_submissions.py

Output (Step 1):
  a2_validation_TIMESTAMP.txt

Output (Step 2):
  a2_games/
    001.html
    002.html
    ...
  a2_warnings_TIMESTAMP.txt
"""

import requests
import os
import re
from datetime import datetime
from openpyxl import load_workbook

# ── Pre-filled config ─────────────────────────────────────────────────────────
CANVAS_BASE_URL = "https://jhu.instructure.com"
API_TOKEN       = ""        # Leave blank to be prompted
COURSE_ID       = "125978"
ASSIGNMENT_ID   = "1290056"
OUTPUT_DIR      = "a2_games"
# ─────────────────────────────────────────────────────────────────────────────

SKIP_ROLES = {"TeacherEnrollment", "TaEnrollment", "DesignerEnrollment"}


def prompt(message):
    return input(message).strip()


def fetch_all(url, headers, params=None):
    results = []
    params = {**(params or {}), "per_page": 100}
    while url:
        resp = requests.get(url, headers=headers, params=params)
        resp.raise_for_status()
        results.extend(resp.json())
        url = None
        for part in resp.headers.get("Link", "").split(","):
            if 'rel="next"' in part:
                url = part[part.index("<") + 1: part.index(">")]
                params = {}
                break
    return results


def get_enrollment_roles(base_url, headers, course_id):
    url = f"{base_url}/api/v1/courses/{course_id}/enrollments"
    enrollments = fetch_all(url, headers)
    return {str(e["user_id"]): e["type"] for e in enrollments}


def load_table1(xlsx_path):
    """
    Read Public_Assigned_Names.xlsx.
    Returns dict: {jhed (lowercase): submission_id (lowercase)}
    Columns: Student Name, JHED, SubmissionID, HTML File, Title File, Prompt File
    """
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    table1 = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        _, jhed, submission_id, _, _, _ = row
        if jhed and submission_id:
            table1[str(jhed).strip().lower()] = str(submission_id).strip().lower()
    wb.close()
    return table1


def load_table2(xlsx_path):
    """
    Read PRIVATE_Mapping.xlsx.
    Returns dict: {submission_id (lowercase): anon_id (e.g. "003")}
    Columns: Student Name, JHED, SubmissionID, Anonymous ID, Anonymous File
    """
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    mapping = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        _, _, submission_id, anon_id, _ = row
        if submission_id and anon_id:
            mapping[str(submission_id).strip().lower()] = str(anon_id).strip().zfill(3)
    wb.close()
    return mapping


def get_submissions(base_url, headers, course_id, assignment_id):
    url = (
        f"{base_url}/api/v1/courses/{course_id}"
        f"/assignments/{assignment_id}/submissions"
    )
    return fetch_all(url, headers, {"include[]": "attachments"})


def extract_base_name(filename):
    """Strip Canvas auto-suffix: 'xaji0y-3.html' -> ('xaji0y', '.html')"""
    name, ext = os.path.splitext(filename)
    name = re.sub(r'-\d+$', '', name)
    return name.lower(), ext.lower()


def get_latest_by_ext(submission, ext):
    """Return the most recently submitted attachment with given extension, or None."""
    attachments = submission.get("attachments", [])
    matched = [
        a for a in attachments
        if os.path.splitext(a.get("filename", ""))[1].lower() == ext
    ]
    if not matched:
        return None
    matched.sort(key=lambda a: a.get("created_at", ""), reverse=True)
    return matched[0]


def download_file(url, headers, dest_path):
    resp = requests.get(url, headers=headers)
    resp.raise_for_status()
    with open(dest_path, "wb") as f:
        f.write(resp.content)


# ── Step 1: Validate filenames ────────────────────────────────────────────────
def step1_validate(base_url, token, course_id, assignment_id, table1):
    headers = {"Authorization": f"Bearer {token}"}
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    print("\nFetching course enrollment data...")
    role_map = get_enrollment_roles(base_url, headers, course_id)

    print("Fetching submissions...")
    submissions = get_submissions(base_url, headers, course_id, assignment_id)

    issues = []
    clean  = []

    for sub in submissions:
        uid  = str(sub.get("user_id", ""))
        role = role_map.get(uid, "Unknown")
        if role in SKIP_ROLES:
            continue
        if sub.get("workflow_state") == "unsubmitted":
            display_name = sub.get("user", {}).get("name", f"User {uid}")
            issues.append({
                "display_name": display_name,
                "uid": uid,
                "jhed": "",
                "expected_submission_id": "",
                "problems": ["No submission found"],
            })
            continue

        display_name = sub.get("user", {}).get("name", f"User {uid}")
        attachments  = sub.get("attachments", [])

        # Build map of base_name -> {ext -> original filename}
        submitted_bases = {}
        for a in attachments:
            fname = a.get("filename", "")
            base, ext = extract_base_name(fname)
            submitted_bases.setdefault(base, {})[ext] = fname

        # Resolve expected SubmissionID via JHED
        sis_login = sub.get("user", {}).get("login_id", "")
        jhed = sis_login.replace("@jh.edu", "").lower() if sis_login else ""
        expected_submission_id = table1.get(jhed)

        problems = []

        if not expected_submission_id:
            problems.append(
                f"Could not find SubmissionID for JHED '{jhed}' in Public_Assigned_Names.xlsx"
            )
        else:
            # Required: {submission_id}.html
            if expected_submission_id not in submitted_bases or \
               ".html" not in submitted_bases[expected_submission_id]:
                problems.append(
                    f"Missing or incorrectly named HTML file "
                    f"(expected: {expected_submission_id}.html)"
                )

            # Required: {submission_id}_prompt.txt
            prompt_base = expected_submission_id + "_prompt"
            if prompt_base not in submitted_bases or \
               ".txt" not in submitted_bases[prompt_base]:
                problems.append(
                    f"Missing or incorrectly named prompt file "
                    f"(expected: {expected_submission_id}_prompt.txt)"
                )

            # Optional: {submission_id}_title.txt
            title_base = expected_submission_id + "_title"
            if title_base not in submitted_bases or \
               ".txt" not in submitted_bases[title_base]:
                problems.append(
                    f"[Optional] Title file not submitted "
                    f"(expected: {expected_submission_id}_title.txt)"
                )

        if problems:
            issues.append({
                "display_name":    display_name,
                "uid":             uid,
                "jhed":            jhed,
                "expected_submission_id": expected_submission_id or "unknown",
                "problems":        problems,
            })
        else:
            clean.append(display_name)

    # Write report
    report_path = f"a2_validation_{timestamp}.txt"
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(f"A2 Filename Validation Report — {timestamp}\n")
        f.write("=" * 60 + "\n\n")
        f.write(f"Clean submissions : {len(clean)}\n")
        f.write(f"Issues found      : {len(issues)}\n\n")
        if issues:
            f.write("STUDENTS TO REACH OUT TO:\n")
            f.write("-" * 40 + "\n")
            for item in issues:
                f.write(
                    f"Student:  {item['display_name']} "
                    f"(JHED: {item['jhed']}  SubmissionID: {item["expected_submission_id"]})\n"
                )
                for p in item["problems"]:
                    prefix = "  [optional]  " if "[Optional]" in p else "  ⚠️   "
                    f.write(f"{prefix}{p}\n")
                f.write("\n")
        else:
            f.write("All submissions passed filename validation.\n")

    required_issues = [
        i for i in issues
        if any("[Optional]" not in p for p in i["problems"])
    ]

    print(f"\n{'=' * 50}")
    print(f"Validation complete!")
    print(f"  Clean:   {len(clean)} students")
    print(f"  Issues:  {len(issues)} students")
    print(f"  Report:  {report_path}")
    if required_issues:
        print(f"\n  ⚠️  {len(required_issues)} student(s) need to fix required files.")
    print(f"{'=' * 50}")


# ── Step 2: Download and anonymize ────────────────────────────────────────────
def step2_download(base_url, token, course_id, assignment_id, table2):
    headers   = {"Authorization": f"Bearer {token}"}
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    print("\nFetching course enrollment data...")
    role_map = get_enrollment_roles(base_url, headers, course_id)

    print("Fetching submissions...")
    submissions = get_submissions(base_url, headers, course_id, assignment_id)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    warnings      = []
    success_count = 0

    for sub in submissions:
        uid  = str(sub.get("user_id", ""))
        role = role_map.get(uid, "Unknown")
        if role in SKIP_ROLES:
            continue
        if sub.get("workflow_state") == "unsubmitted":
            continue

        display_name = sub.get("user", {}).get("name", f"User {uid}")
        attachment   = get_latest_by_ext(sub, ".html")

        if not attachment:
            warnings.append({
                "display_name":     display_name,
                "uid":              uid,
                "original_filename": "N/A",
                "issue":            "No HTML file found in submission",
            })
            continue

        original_filename = attachment.get("filename", "")
        base_name, _      = extract_base_name(original_filename)

        if base_name not in table2:
            warnings.append({
                "display_name":     display_name,
                "uid":              uid,
                "original_filename": original_filename,
                "issue": (
                    f"SubmissionID '{base_name}' not found in PRIVATE_Mapping.xlsx. "
                    "Student may have used wrong filename."
                ),
            })
            continue

        anon_id   = table2[base_name]
        dest_path = os.path.join(OUTPUT_DIR, f"{anon_id}.html")

        print(f"  {display_name}: {original_filename} → {anon_id}.html")
        download_file(attachment["url"], headers, dest_path)
        success_count += 1

    # Write warnings
    warnings_path = f"a2_warnings_{timestamp}.txt"
    with open(warnings_path, "w", encoding="utf-8") as f:
        if warnings:
            f.write(f"A2 Download Warnings — {timestamp}\n")
            f.write("=" * 60 + "\n\n")
            for w in warnings:
                f.write(f"Student:  {w['display_name']} (User ID: {w['uid']})\n")
                f.write(f"File:     {w['original_filename']}\n")
                f.write(f"Issue:    {w['issue']}\n")
                f.write("-" * 40 + "\n")
        else:
            f.write("No warnings — all HTML files downloaded successfully.\n")

    print(f"\n{'=' * 50}")
    print(f"Done!")
    print(f"  Files saved to:  ./{OUTPUT_DIR}/")
    print(f"  Warnings:        {warnings_path}")
    print(f"  Downloaded:      {success_count} files")
    if warnings:
        print(f"  ⚠️  {len(warnings)} issue(s) — see warnings file")
    print(f"{'=' * 50}")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    global API_TOKEN

    print("=" * 50)
    print("  A2 Game Submission Anonymizer")
    print("=" * 50)
    print(f"  Canvas:        {CANVAS_BASE_URL}")
    print(f"  Course ID:     {COURSE_ID}")
    print(f"  Assignment ID: {ASSIGNMENT_ID}")

    if not API_TOKEN:
        API_TOKEN = prompt("\nEnter your Canvas API token: ")

    print("\nWhat would you like to do?")
    print("  1 — Step 1: Validate filenames only (no download)")
    print("  2 — Step 2: Download and anonymize HTML files")
    choice = prompt("Enter 1 / 2: ")

    if choice == "1":
        xlsx_path = prompt(
            "Path to Public_Assigned_Names.xlsx [default: Public_Assigned_Names.xlsx]: "
        )
        if not xlsx_path:
            xlsx_path = "Public_Assigned_Names.xlsx"
        print(f"\nLoading SubmissionID mapping from {xlsx_path}...")
        table1 = load_table1(xlsx_path)
        print(f"  Loaded {len(table1)} entries.")
        step1_validate(CANVAS_BASE_URL, API_TOKEN, COURSE_ID, ASSIGNMENT_ID, table1)

    elif choice == "2":
        xlsx_path = prompt(
            "Path to PRIVATE_Mapping.xlsx [default: PRIVATE_Mapping.xlsx]: "
        )
        if not xlsx_path:
            xlsx_path = "PRIVATE_Mapping.xlsx"
        print(f"\nLoading anonymous mapping from {xlsx_path}...")
        table2 = load_table2(xlsx_path)
        print(f"  Loaded {len(table2)} entries.")
        step2_download(CANVAS_BASE_URL, API_TOKEN, COURSE_ID, ASSIGNMENT_ID, table2)

    else:
        print("Invalid choice. Please enter 1 or 2.")


if __name__ == "__main__":
    main()
