"""
A3 Job Hunting Submission Processor
=====================================
Run this script after each attempt's deadline.
You will be prompted to select which attempt to process (1, 2, or 3).

Output per run:
  a3_submissions/
    Attempt1/
      001/
        mask-Attempt1.docx
      002/
        ...
  a3_mapping_Attempt1_TIMESTAMP.xlsx  -- private, one sheet per Job ID
  a3_warnings_Attempt1_TIMESTAMP.txt  -- students with invalid filenames

Usage:
  pip install requests openpyxl
  python a3_process_submissions.py
"""

import requests
import os
import re
from datetime import datetime
from openpyxl import Workbook

# ── Pre-filled config ─────────────────────────────────────────────────────────
CANVAS_BASE_URL = "https://jhu.instructure.com"
API_TOKEN       = ""       # Leave blank to be prompted
COURSE_ID       = "125978"

ATTEMPTS = {
    "1": {"label": "Attempt1", "assignment_id": "1290088"},
    "2": {"label": "Attempt2", "assignment_id": "1290101"},
    "3": {"label": "Attempt3", "assignment_id": "1290112"},
}

OUTPUT_BASE_DIR = "a3_submissions"
# ─────────────────────────────────────────────────────────────────────────────

SKIP_ROLES = {"TeacherEnrollment", "TaEnrollment", "DesignerEnrollment"}

# Expected: JHED_JobID_mask.docx (with optional Canvas suffix -1, -2, -3)
FILENAME_PATTERN = re.compile(
    r'^(?P<jhed>[a-zA-Z0-9]+)_(?P<job_id>\d{3})_(?P<mask>[a-zA-Z0-9]+)(?:-\d+)?\.docx$',
    re.IGNORECASE
)


def prompt(message):
    return input(message).strip()


def fetch_all(url, headers, params=None):
    """Fetch all pages from a paginated Canvas API endpoint."""
    results = []
    params = {**(params or {}), "per_page": 100}
    while url:
        resp = requests.get(url, headers=headers, params=params)
        resp.raise_for_status()
        results.extend(resp.json())
        url = None
        for part in resp.headers.get("Link", "").split(","):
            if 'rel="next"' in part:
                url = part[part.index("<") + 1: part.index(">")]
                params = {}
                break
    return results


def get_enrollment_roles(base_url, headers, course_id):
    url = f"{base_url}/api/v1/courses/{course_id}/enrollments"
    enrollments = fetch_all(url, headers)
    return {str(e["user_id"]): e["type"] for e in enrollments}


def get_latest_docx(submission):
    """Return the most recently submitted .docx attachment, or None."""
    attachments = submission.get("attachments", [])
    docx_files = [
        a for a in attachments
        if a.get("filename", "").lower().endswith(".docx")
    ]
    if not docx_files:
        return None
    docx_files.sort(key=lambda a: a.get("created_at", ""), reverse=True)
    return docx_files[0]


def download_file(url, headers, dest_path):
    resp = requests.get(url, headers=headers)
    resp.raise_for_status()
    with open(dest_path, "wb") as f:
        f.write(resp.content)


def run(base_url, token, course_id, attempt_label, assignment_id):
    headers = {"Authorization": f"Bearer {token}"}
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    print("\nFetching course enrollment data...")
    role_map = get_enrollment_roles(base_url, headers, course_id)

    print(f"Fetching submissions for {attempt_label}...")
    url = (
        f"{base_url}/api/v1/courses/{course_id}"
        f"/assignments/{assignment_id}/submissions"
    )
    submissions = fetch_all(url, headers, {"include[]": "attachments"})

    # {job_id: [row_dict, ...]}
    mapping_by_job = {}
    warnings = []

    for sub in submissions:
        uid = str(sub.get("user_id", ""))
        role = role_map.get(uid, "Unknown")
        if role in SKIP_ROLES:
            continue
        if sub.get("workflow_state") == "unsubmitted":
            continue

        display_name = sub.get("user", {}).get("name", f"User {uid}")
        attachment = get_latest_docx(sub)

        if not attachment:
            warnings.append({
                "user_id": uid,
                "display_name": display_name,
                "original_filename": "N/A",
                "issue": "No .docx file found in submission",
            })
            continue

        original_filename = attachment.get("filename", "")
        match = FILENAME_PATTERN.match(original_filename)

        if not match:
            warnings.append({
                "user_id": uid,
                "display_name": display_name,
                "original_filename": original_filename,
                "issue": "Filename does not match required format: JHED_JobID_mask.docx",
            })
            continue

        jhed   = match.group("jhed")
        job_id = match.group("job_id")
        mask   = match.group("mask")

        # Download to a3_submissions/Attempt1/001/mask-Attempt1.docx
        job_folder = os.path.join(OUTPUT_BASE_DIR, attempt_label, job_id)
        os.makedirs(job_folder, exist_ok=True)
        clean_filename = f"{mask}-{attempt_label}.docx"
        dest_path = os.path.join(job_folder, clean_filename)

        print(f"  {display_name} → {attempt_label}/{job_id}/{clean_filename}")
        download_file(attachment["url"], headers, dest_path)

        row = {
            "Anonymous Name":    f"{mask}-{attempt_label}",
            "Mask":              mask,
            "JHED":              jhed,
            "Job ID":            job_id,
            "Student Name":      display_name,
            "Canvas User ID":    uid,
            "Original Filename": original_filename,
            "Submitted At":      sub.get("submitted_at", ""),
        }
        mapping_by_job.setdefault(job_id, []).append(row)

    # Write mapping Excel — one sheet per Job ID
    mapping_path = f"a3_mapping_{attempt_label}_{timestamp}.xlsx"
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    headers_row = [
        "Anonymous Name", "Mask", "JHED", "Job ID",
        "Student Name", "Canvas User ID", "Original Filename", "Submitted At",
    ]

    if mapping_by_job:
        for job_id in sorted(mapping_by_job.keys()):
            ws = wb.create_sheet(title=f"Job {job_id}")
            ws.append(headers_row)
            for row in mapping_by_job[job_id]:
                ws.append([row[h] for h in headers_row])
    else:
        # Create a blank sheet so the file is valid
        ws = wb.create_sheet(title="No Submissions")
        ws.append(headers_row)

    wb.save(mapping_path)

    # Write warnings file
    warnings_path = f"a3_warnings_{attempt_label}_{timestamp}.txt"
    with open(warnings_path, "w", encoding="utf-8") as f:
        if warnings:
            f.write(f"A3 Warnings — {attempt_label} — {timestamp}\n")
            f.write("=" * 60 + "\n\n")
            for w in warnings:
                f.write(f"Student:  {w['display_name']} (User ID: {w['user_id']})\n")
                f.write(f"File:     {w['original_filename']}\n")
                f.write(f"Issue:    {w['issue']}\n")
                f.write("-" * 40 + "\n")
        else:
            f.write(f"No warnings — all {attempt_label} submissions passed format validation.\n")

    # Summary
    total = sum(len(v) for v in mapping_by_job.values())
    print(f"\n{'=' * 50}")
    print(f"Done!")
    print(f"  Files saved to:   ./{OUTPUT_BASE_DIR}/{attempt_label}/")
    print(f"  Mapping saved to: {mapping_path}  (PRIVATE)")
    print(f"  Warnings:         {warnings_path}")
    print(f"  Valid submissions: {total}")
    if warnings:
        print(f"  ⚠️  {len(warnings)} submission(s) need attention — see warnings file")
    print(f"{'=' * 50}")


def main():
    global API_TOKEN

    print("=" * 50)
    print("  A3 Job Hunting Submission Processor")
    print("=" * 50)
    print(f"  Canvas:    {CANVAS_BASE_URL}")
    print(f"  Course ID: {COURSE_ID}")

    if not API_TOKEN:
        API_TOKEN = prompt("\nEnter your Canvas API token: ")

    print("\nWhich attempt would you like to process?")
    print("  1 — Attempt 1")
    print("  2 — Attempt 2")
    print("  3 — Attempt 3")
    choice = prompt("Enter 1 / 2 / 3: ")

    if choice not in ATTEMPTS:
        print("Invalid choice. Please enter 1, 2, or 3.")
        return

    attempt_label  = ATTEMPTS[choice]["label"]
    assignment_id  = ATTEMPTS[choice]["assignment_id"]

    print(f"\nProcessing {attempt_label} (Assignment ID: {assignment_id})...")
    run(CANVAS_BASE_URL, API_TOKEN, COURSE_ID, attempt_label, assignment_id)


if __name__ == "__main__":
    main()
