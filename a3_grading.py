"""
A3 Job Hunting Grader
======================
Scrapes the professor's leaderboard website to get Mean Ranks,
matches results to students via the a3_mapping files,
and calculates final scores using the grading formula.

Grading Formula:
  - Ranked 1st:         17 points
  - Other ranks:        17 - (r - 1) * (9 / (n - 1))
    where r = best Mean Rank, n = total candidates for that job
  - No submission /
    format violation:   0 points

Usage:
  pip install requests openpyxl beautifulsoup4
  python a3_grading.py

You will be prompted for:
  - The leaderboard URL (e.g. https://www.bigdataist.com/job51/)
  - The a3_mapping Excel files (one per attempt)

Output:
  a3_grades_TIMESTAMP.xlsx
    Sheet 1: Full Grading Table (one row per student)
    Sheet 2: Leaderboard Data (raw scraped data for reference)
"""

import requests
import re
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup

# ── Pre-filled config ─────────────────────────────────────────────────────────
CANVAS_BASE_URL = "https://jhu.instructure.com"
COURSE_ID       = "125978"
TOTAL_POINTS    = 17
# ─────────────────────────────────────────────────────────────────────────────


def prompt(message):
    return input(message).strip()


def fetch_html(url):
    resp = requests.get(url, timeout=10)
    resp.raise_for_status()
    return BeautifulSoup(resp.text, "html.parser")


def scrape_leaderboard(base_url):
    """
    Scrape all job leaderboard pages.
    Returns:
      job_info:    {job_id: {"title": str, "n": int}}
      rankings:    {job_id: {mask_attempt: mean_rank}}
                   e.g. {"004": {"tienan-Attempt2": 1.2, ...}}
    """
    print(f"\nFetching main leaderboard: {base_url}")
    soup = fetch_html(base_url)

    job_info = {}
    rankings = {}

    # Find all links to individual job leaderboard pages
    links = soup.find_all("a", href=re.compile(r"job_\d{3}_leaderboard_student\.html"))
    if not links:
        print("  Warning: No job leaderboard links found on main page.")
        return job_info, rankings

    for link in links:
        href = link.get("href", "")
        job_id_match = re.search(r"job_(\d{3})_leaderboard_student\.html", href)
        if not job_id_match:
            continue
        job_id = job_id_match.group(1)

        # Build absolute URL
        if href.startswith("http"):
            job_url = href
        else:
            job_url = base_url.rstrip("/") + "/" + href.lstrip("/")

        print(f"  Scraping Job {job_id}: {job_url}")
        try:
            job_soup = fetch_html(job_url)
        except Exception as e:
            print(f"    Warning: Could not fetch {job_url} — {e}")
            continue

        # Extract job title and total candidates from page text
        page_text = job_soup.get_text()
        title_match = re.search(r"(?:TITLE|COMPANY):\s*(.+?)(?:\n|Job ID)", page_text)
        title = title_match.group(1).strip() if title_match else f"Job {job_id}"

        n_match = re.search(r"Total Candidates:\s*(\d+)", page_text)
        n = int(n_match.group(1)) if n_match else 0

        job_info[job_id] = {"title": title, "n": n}
        rankings[job_id] = {}

        # Extract each candidate's Mean Rank
        # Pattern in page: "#N  mask-AttemptX ... Mean Rank N.N"
        # Use regex on raw text for robustness
        blocks = re.findall(
            r'(?:#{1,2}\d+)\s+([a-zA-Z0-9_\-]+-Attempt\d)\s.*?Mean\s+Rank\s+([\d.]+)',
            page_text,
            re.DOTALL
        )
        for mask_attempt, mean_rank_str in blocks:
            mask_attempt = mask_attempt.strip()
            try:
                mean_rank = float(mean_rank_str)
            except ValueError:
                continue
            rankings[job_id][mask_attempt] = mean_rank

        print(f"    Title: {title} | Candidates: {n} | Entries scraped: {len(rankings[job_id])}")

    return job_info, rankings


def load_mapping_files(xlsx_paths):
    """
    Load one or more a3_mapping Excel files.
    Returns dict: {mask_attempt: {"student_name": ..., "jhed": ..., "job_id": ..., "attempt": ...}}
    e.g. {"tienan-Attempt2": {"student_name": "Zhang, Tiange", "jhed": "tzhang98", ...}}
    """
    combined = {}
    for path in xlsx_paths:
        if not os.path.exists(path):
            print(f"  Warning: Mapping file not found: {path}")
            continue
        wb = load_workbook(path, read_only=True)
        for sheet in wb.sheetnames:
            if not sheet.startswith("Job"):
                continue
            ws = wb[sheet]
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue  # skip header
                # Columns: Anonymous Name, Mask, JHED, Job ID, Student Name,
                #          Canvas User ID, Original Filename, Submitted At
                anon_name, mask, jhed, job_id, student_name = row[0], row[1], row[2], row[3], row[4]
                if not anon_name:
                    continue
                attempt_match = re.search(r'(Attempt\d)$', str(anon_name))
                attempt = attempt_match.group(1) if attempt_match else ""
                key = f"{mask}-{attempt}"
                combined[key] = {
                    "student_name": student_name,
                    "jhed":         jhed,
                    "job_id":       str(job_id).zfill(3) if job_id else "",
                    "attempt":      attempt,
                    "mask":         mask,
                }
        wb.close()
    return combined


def calculate_score(r, n):
    """
    r = Mean Rank (float), n = total candidates (int)
    Returns score out of 17, rounded to 2 decimal places.
    """
    if n <= 1:
        return TOTAL_POINTS  # only candidate, full marks
    score = TOTAL_POINTS - (r - 1) * (9 / (n - 1))
    return round(max(score, 0), 2)


def run(leaderboard_url, mapping_paths, roster):
    """
    roster: {jhed: student_name} from Canvas enrollment (optional, for completeness check)
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # 1. Scrape leaderboard
    job_info, rankings = scrape_leaderboard(leaderboard_url)

    # 2. Load mapping files
    print("\nLoading mapping files...")
    mapping = load_mapping_files(mapping_paths)
    print(f"  Loaded {len(mapping)} submission entries.")

    # 3. Build per-student data
    # Group by student (jhed)
    students = {}  # jhed -> {attempt: {job_id, mask, rank, n, score}}
    for key, info in mapping.items():
        jhed   = info["jhed"]
        attempt = info["attempt"]
        job_id  = info["job_id"]
        mask    = info["mask"]

        if jhed not in students:
            students[jhed] = {
                "student_name": info["student_name"],
                "attempts": {}
            }

        # Look up rank from leaderboard
        mask_attempt = f"{mask}-{attempt}"
        mean_rank = None
        if job_id in rankings and mask_attempt in rankings[job_id]:
            mean_rank = rankings[job_id][mask_attempt]

        n = job_info.get(job_id, {}).get("n", 0)
        score = calculate_score(mean_rank, n) if mean_rank is not None else None

        students[jhed]["attempts"][attempt] = {
            "job_id":    job_id,
            "mask":      mask,
            "mean_rank": mean_rank,
            "n":         n,
            "score":     score,
        }

    # 4. Calculate best score per student
    rows = []
    for jhed, data in sorted(students.items()):
        name    = data["student_name"]
        attempts = data["attempts"]

        a1 = attempts.get("Attempt1", {})
        a2 = attempts.get("Attempt2", {})
        a3 = attempts.get("Attempt3", {})

        # Best score = highest score across all attempts
        all_scores = [
            a["score"] for a in [a1, a2, a3]
            if a and a.get("score") is not None
        ]
        best_score = max(all_scores) if all_scores else 0

        # Find which attempt gave best score
        best_attempt = ""
        best_rank    = ""
        best_job     = ""
        best_n       = ""
        for label, a in [("Attempt1", a1), ("Attempt2", a2), ("Attempt3", a3)]:
            if a and a.get("score") == best_score:
                best_attempt = label
                best_rank    = a.get("mean_rank", "")
                best_job     = a.get("job_id", "")
                best_n       = a.get("n", "")
                break

        rows.append({
            "Student Name":      name,
            "JHED":              jhed,
            "A1 Job":            a1.get("job_id", ""),
            "A1 Mean Rank":      a1.get("mean_rank", "N/A") if a1 else "No submission",
            "A1 n":              a1.get("n", "") if a1 else "",
            "A1 Score":          a1.get("score", 0) if a1 else 0,
            "A2 Job":            a2.get("job_id", ""),
            "A2 Mean Rank":      a2.get("mean_rank", "N/A") if a2 else "No submission",
            "A2 n":              a2.get("n", "") if a2 else "",
            "A2 Score":          a2.get("score", 0) if a2 else 0,
            "A3 Job":            a3.get("job_id", ""),
            "A3 Mean Rank":      a3.get("mean_rank", "N/A") if a3 else "No submission",
            "A3 n":              a3.get("n", "") if a3 else "",
            "A3 Score":          a3.get("score", 0) if a3 else 0,
            "Best Attempt":      best_attempt,
            "Best Job":          best_job,
            "Best Mean Rank":    best_rank,
            "Best n":            best_n,
            "Final Score (/17)": best_score,
        })

    # 5. Write Excel output
    output_path = f"a3_grades_{timestamp}.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: Grading table
    ws1 = wb.create_sheet("Grades")
    headers = [
        "Student Name", "JHED",
        "A1 Job", "A1 Mean Rank", "A1 n", "A1 Score",
        "A2 Job", "A2 Mean Rank", "A2 n", "A2 Score",
        "A3 Job", "A3 Mean Rank", "A3 n", "A3 Score",
        "Best Attempt", "Best Job", "Best Mean Rank", "Best n",
        "Final Score (/17)",
    ]
    ws1.append(headers)
    for row in sorted(rows, key=lambda r: r["Student Name"]):
        ws1.append([row[h] for h in headers])

    # Sheet 2: Raw leaderboard data
    ws2 = wb.create_sheet("Leaderboard Data")
    ws2.append(["Job ID", "Job Title", "Total Candidates", "Mask-Attempt", "Mean Rank"])
    for job_id in sorted(rankings.keys()):
        title = job_info.get(job_id, {}).get("title", "")
        n     = job_info.get(job_id, {}).get("n", "")
        for mask_attempt, mean_rank in sorted(
            rankings[job_id].items(), key=lambda x: x[1]
        ):
            ws2.append([job_id, title, n, mask_attempt, mean_rank])

    wb.save(output_path)

    print(f"\n{'=' * 50}")
    print(f"Done!")
    print(f"  Grades saved to: {output_path}")
    print(f"  Students graded: {len(rows)}")
    print(f"{'=' * 50}")


def main():
    print("=" * 50)
    print("  A3 Job Hunting Grader")
    print("=" * 50)

    leaderboard_url = prompt(
        "\nLeaderboard URL (e.g. https://www.bigdataist.com/jobXX/): "
    )

    print("\nEnter paths to a3_mapping Excel files (one per attempt processed).")
    print("Press Enter with no input when done.")
    mapping_paths = []
    i = 1
    while True:
        p = prompt(f"  Mapping file {i} path (or press Enter to finish): ")
        if not p:
            break
        mapping_paths.append(p)
        i += 1

    if not mapping_paths:
        print("No mapping files provided. Exiting.")
        return

    run(leaderboard_url, mapping_paths, roster={})


if __name__ == "__main__":
    main()
